from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import pandas as pd
import re
import os
import uuid
import tempfile
from typing import Dict
from typing import Tuple

app = FastAPI(title="Markdown to Excel Converter")


class MarkdownInput(BaseModel):
    content: str


# 存储文件信息
file_storage: Dict[str, str] = {}


def extract_markdown_table(text: str) -> str:
    """提取markdown表格内容"""
    # 如果文本直接以### 测试用例开头，直接处理
    if text.strip().startswith('### 测试用例') or text.strip().startswith('###测试用例'):
        # 找到表格开始的位置（第一个|字符的行）
        lines = text.strip().split('\n')
        table_start = -1

        for i, line in enumerate(lines):
            if '|' in line and line.strip().startswith('|'):
                table_start = i
                break

        if table_start == -1:
            raise ValueError("未找到表格内容")

        # 返回表格部分
        return '\n'.join(lines[table_start:]).strip()

    # 原有的查找逻辑作为备用
    pattern = r'###\s*测试用例\s*\n(.*?)(?=\n###|\Z)'
    match = re.search(pattern, text, re.DOTALL)

    if not match:
        raise ValueError("未找到'### 测试用例'标记或其后的内容")

    content = match.group(1).strip()

    # 确保内容包含表格
    if '|' not in content:
        raise ValueError("未找到表格内容")

    return content


def parse_markdown_table(markdown_table: str) -> pd.DataFrame:
    """解析markdown表格为DataFrame"""
    lines = [line.strip() for line in markdown_table.strip().split('\n') if line.strip()]

    if len(lines) < 2:
        raise ValueError("表格格式不正确，至少需要表头和分隔符行")

    # 找到表头行（第一行包含|的行）
    header_line = None
    data_start_index = 0

    for i, line in enumerate(lines):
        if '|' in line:
            header_line = line
            data_start_index = i
            break

    if not header_line:
        raise ValueError("未找到表格表头")

    # 解析表头
    headers = []
    header_parts = header_line.split('|')
    for part in header_parts:
        part = part.strip()
        if part:  # 跳过空的部分
            headers.append(part)

    if not headers:
        raise ValueError("表头解析失败")

    # 跳过分隔符行，从数据行开始解析
    data_lines = lines[data_start_index + 2:]  # +2 跳过表头和分隔符行

    # 解析数据行
    rows = []
    for line in data_lines:
        if line.strip() and '|' in line:
            cells = []
            # 使用更精确的分割方法
            parts = line.split('|')

            for part in parts:
                part = part.strip()
                if part:  # 只添加非空的部分
                    # 处理HTML标签
                    part = re.sub(r'<br\s*/?>', '\n', part)
                    part = re.sub(r'<[^>]+>', '', part)
                    cells.append(part.strip())

            # 确保单元格数量与表头匹配
            while len(cells) < len(headers):
                cells.append("")

            # 只取前len(headers)个单元格
            rows.append(cells[:len(headers)])

    if not rows:
        raise ValueError("未找到有效的数据行")

    # 创建DataFrame
    df = pd.DataFrame(rows, columns=headers)
    return df


def create_excel_file(df: pd.DataFrame) -> Tuple[str, str]:
    """创建Excel文件并返回文件路径和文件ID"""
    temp_dir = tempfile.gettempdir()
    file_id = uuid.uuid4().hex[:8]
    filename = f"test_cases_{file_id}.xlsx"
    filepath = os.path.join(temp_dir, filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='测试用例', index=False)
        worksheet = writer.sheets['测试用例']

        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass

            # 设置列宽
            adjusted_width = min(max(max_length // 3 + 5, 15), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # 设置样式
        try:
            from openpyxl.styles import Font, PatternFill, Alignment

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # 设置表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 设置数据行样式
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # 设置行高
            for row_num in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row_num].height = 60

        except ImportError:
            pass

    return filepath, file_id


def get_base_url(request: Request) -> str:
    """获取服务器的基础URL"""
    # 优先使用环境变量中设置的公网地址
    base_url = os.getenv('BASE_URL', 'http://120.131.12.60:8000')

    # 如果没有设置环境变量，尝试从请求中获取
    if not os.getenv('BASE_URL'):
        # 检查是否有代理头信息
        forwarded_proto = request.headers.get('x-forwarded-proto', 'http')
        forwarded_host = request.headers.get('x-forwarded-host')

        if forwarded_host:
            base_url = f"{forwarded_proto}://{forwarded_host}"
        else:
            # 使用请求中的主机信息
            host = request.headers.get('host', '120.131.12.60:8000')
            scheme = 'https' if request.url.scheme == 'https' else 'http'
            base_url = f"{scheme}://{host}"

    return base_url.rstrip('/')


@app.post("/convert")
async def convert_to_excel(data: MarkdownInput, request: Request):
    """将markdown内容转换为Excel文件并返回下载URL"""
    try:
        print(f"接收到的内容: {data.content[:200]}...")

        # 提取表格内容
        table_content = extract_markdown_table(data.content)
        print(f"提取的表格内容: {table_content[:200]}...")

        # 解析为DataFrame
        df = parse_markdown_table(table_content)
        print(f"解析得到 {len(df)} 行 {len(df.columns)} 列数据")

        # 创建Excel文件
        filepath, file_id = create_excel_file(df)

        # 存储文件信息
        file_storage[file_id] = filepath

        # 获取基础URL并构建完整的下载链接
        base_url = get_base_url(request)
        download_url = f"{base_url}/download/{file_id}"

        return {
            "success": True,
            "message": "转换成功",
            "download_url": download_url,
            "file_id": file_id,
            "rows_count": len(df),
            "columns_count": len(df.columns),
            "columns": df.columns.tolist(),
            "filename": f"test_cases_{file_id}.xlsx"
        }

    except ValueError as e:
        print(f"ValueError: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        print(f"Exception: {str(e)}")
        raise HTTPException(status_code=500, detail=f"转换失败: {str(e)}")


@app.get("/download/{file_id}")
async def download_file(file_id: str):
    """下载Excel文件"""
    if file_id not in file_storage:
        raise HTTPException(status_code=404, detail="文件不存在")

    filepath = file_storage[file_id]
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="文件不存在")

    from fastapi.responses import FileResponse
    return FileResponse(
        path=filepath,
        filename=f"test_cases_{file_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/")
async def root():
    return {
        "message": "Markdown to Excel Converter API",
        "endpoint": "POST /convert",
        "usage": "发送 markdown 表格内容到 /convert 接口"
    }


if __name__ == "__main__":
    import uvicorn

    # 设置环境变量指定公网地址
    os.environ['BASE_URL'] = 'http://120.131.12.60:8000'
    uvicorn.run(app, host="0.0.0.0", port=8000)
